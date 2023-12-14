'''

A short Python program for retirement planning.
Calculations are done on a yearly basis.

Currently, it supports single filers and married
filing jointly. Feel free to modify for other cases.

Martin-D. Lacasse - 2023

This program comes with no guarantee. Use at your own risks.

'''

######################################################################
# Some of the modules required:
import sys
import datetime
import numpy as np
import math

# Our own modules:
import utils as u
import rates
import tax as tx

######################################################################


def setVerbose(state):
    '''
    Control verbosity of print statements.
    '''
    u.setVerbose(state)


class Plan:
    def __init__(self, YOB, expectancy):
        '''
        Constructor requires two lists the first one containing the
        years of birth of each spouse and the other, life expectancies.
        To be clear: for singles, a list of one entry, for married
        couples, a list of two entries.
        This nformation will determine the size of arrays required
        for the calculations.
        '''
        # Check inputs.
        self.count = len(YOB)
        assert (self.count == len(expectancy))
        assert (0 < self.count and self.count <= 2)

        self.status = ['single', 'married'][self.count-1]

        self.yob = YOB
        self.expectancy = expectancy

        now = datetime.date.today().year
        # Compute both life horizons through a comprehension.
        self.horizons = [expectancy[i] + YOB[i] - now + 1
                         for i in range(self.count)]
        # Add one more year as we are computing values for next year.
        self.maxHorizon = max(self.horizons) + 2

        # Variables starting with a 'y' are tracking yearly values.
        # Initialize variables to track year after year:
        self.yyear = np.array(range(now, now+self.maxHorizon))

        if self.count == 1:
            ages = age(YOB[0])
            self.y2ages = np.array(range(ages, ages+self.maxHorizon))
        elif self.count == 2:
            ages = np.array([age(YOB[0]), age(YOB[1])])
            self.y2ages = np.array([range(ages[0], ages[0]+self.maxHorizon),
                                   range(ages[1], ages[1]+self.maxHorizon)])
            self.y2ages = self.y2ages.transpose()

        # Keep data in [year][who] for now.
        # We'll transpose later if needed when plotting.
        self.y2accounts = {}
        for aType in ['taxable', 'tax-deferred', 'tax-free']:
            self.y2accounts[aType] = np.zeros((self.maxHorizon, self.count))

        self.beneficiary = np.ones((2))

        self.names = None
        self.timeLists = None

        # Default value for split between spouse is auto.
        self.split = 'auto'

        # Target net income during retirement.
        self.target = None
        self.profile = 'flat'

        self.pensionAmount = None
        self.pensionAge = None

        self.ssecAmount = None
        self.ssecAge = None

        self.y2assetRatios = {}
        self.boundsAR = {}
        for aType in ['taxable', 'tax-deferred', 'tax-free']:
            self.y2assetRatios[aType] = \
                    np.zeros((self.maxHorizon, self.count, 4))
            self.boundsAR[aType] = np.zeros((2, self.count, 4))

        # Set default values on bounds.
        for who in range(self.count):
            self.boundsAR['taxable'][0][who][:] = [0, .25, .5, .25]
            self.boundsAR['taxable'][1][who][:] = [0, .25, .5, .25]
            self.boundsAR['tax-deferred'][0][who][:] = [.6, .4, 0, 0]
            self.boundsAR['tax-deferred'][1][who][:] = [.6, .4, 0, 0]
            self.boundsAR['tax-free'][0][who][:] = [.6, .4, 0, 0]
            self.boundsAR['tax-free'][1][who][:] = [.6, .4, 0, 0]

        self.interpolateAR()

        self.y2source = None
        self.yincome = None

        self.rates = None
        self.rateMethod = 'default'
        self.rateFrm = None
        self.rateTo = None
        self.setRates('default')

        # Track if run ws successful
        self.success = True

    def setInitialAR(self, *, taxableAR, taxDeferredAR, taxFreeAR):
        '''
        Set values of asset distribution in each of the accounts
        for the first horizon year.
        '''
        self.setAR(taxableAR, taxDeferredAR, taxFreeAR, 0)

    def setFinalAR(self, *, taxableAR, taxDeferredAR, taxFreeAR):
        '''
        Set values of asset distribution in each of the accounts
        for the last horizon year.
        '''
        self.setAR(taxableAR, taxDeferredAR, taxFreeAR, 1)

    def setAR(self, taxableAR, taxDeferredAR, taxFreeAR, i):
        '''
        Utility function for setting initial time and final time values
        on asset ratios.
        '''
        assert len(taxableAR) == self.count
        assert len(taxDeferredAR) == self.count
        assert len(taxFreeAR) == self.count

        for i in range(self.count):
            assert len(taxableAR[i]) == 4
            assert abs(sum(taxableAR[i]) - 100) < 0.01

            assert len(taxDeferredAR[i]) == 4
            assert abs(sum(taxDeferredAR[i]) - 100) < 0.01

            assert len(taxFreeAR[i]) == 4
            assert abs(sum(taxFreeAR[i]) - 100) < 0.01

        which = ['Initial', 'Final'][i]

        u.vprint(which, 'asset ratios set to: (%)\n', taxableAR,
                 '\n', taxDeferredAR, '\n', taxFreeAR)
        self.boundsAR['taxable'][i] = np.array(taxableAR)/100
        self.boundsAR['tax-deferred'][i] = np.array(taxDeferredAR)/100
        self.boundsAR['tax-free'][i] = np.array(taxFreeAR)/100

    def interpolateAR(self, method='linear'):
        '''
        Interpolate asset ratio from initial value (today) to
        final value (at the end of horizon.
        '''
        if method == 'linear':
            for accType in ['taxable', 'tax-deferred', 'tax-free']:
                for who in range(self.count):
                    for j in range(4):
                        dat = np.linspace(self.boundsAR[accType][0][who][j],
                                          self.boundsAR[accType][1][who][j],
                                          self.horizons[who])
                        for k in range(self.horizons[who]):
                            self.y2assetRatios[accType][k][who][j] = dat[k]
        else:
            u.xprint('Method', method, 'not supported')

        u.vprint('Interpolated asset ratios using', method, 'method.')

    def setRates(self, method, frm=None, to=None, offset=0):
        '''
        Generate rates for return and inflation based on the method and
        years selected. Optional offsest can be provided to investigate
        effects of return sequence.
        '''
        dr = rates.rates()
        dr.setMethod(method, frm, to)
        self.rateMethod = method
        self.rateFrm = frm
        self.rateTo = to
        # Beware as some inflation calculations cannot start in 2017
        self.rates = dr.genSeries(offset, self.maxHorizon)
        # u.vprint('Generated rate series of', len(self.rates))

    def setAssetBalances(self, *, taxable, taxDeferred, taxFree, beneficiary):
        '''
        Four entries must be provided. The first three are lists
        containing the balance of all assets in each category for
        each spouse. The last one is the fraction of assets left to
        the other spouse as a beneficiary. For single individuals,
        these lists will contain only one entry and the beneficiary
        value is not relevant.
        '''
        assert (len(taxable) == self.count)
        assert (len(taxDeferred) == self.count)
        assert (len(taxFree) == self.count)
        assert (len(beneficiary) == self.count)

        self.y2accounts['taxable'][0][:] = taxable
        self.y2accounts['tax-deferred'][0][:] = taxDeferred
        self.y2accounts['tax-free'][0][:] = taxFree
        self.beneficiary = beneficiary

    # Asset ratios are lists with 3 values: stock, bonds, and fixed assets.
    def setAssetRatiosI(self, *, taxableR, taxDeferredR, taxFreeR):
        '''
        Provide asset ratios between stock, bonds, and fixed assets
        for each asset category and each spouse. For a married couple
        with a 60/40 tax-deferred portfolio, 80/20 tax-free, and
        taxable account with 50% bonds and 50% in fixed assets would
        be as follows:

        plan = retirement.Plan(...)

        plan.setAssetRatiosI(
            taxableR = [[0, 50, 50], [0, 50, 50]],
            taxDeferredR = [[60, 40, 0], [60, 40, 0]],
            taxFreeR = [[80, 20, 0], [80, 20, 0]]
            )

        For single individuals, the list contains only one entry
        with three values. Each triplet values must add up to 100%.
        '''
        self.testAssetRatios(self, taxableR, taxDeferredR, taxFreeR)

        # Convert from percent to decimal.
        self.taxableR[0] = np.array(taxableR)/100
        self.taxDeferredR[0] = np.array(taxDeferredR)/100
        self.taxFreeR[0] = np.array(taxFreeR)/100

    def testAssetRatios(self, taxableR, taxDeferredR, taxFreeR):
        '''
        Determine if entries are correct.
        '''
        assert len(taxableR) == self.count
        assert len(taxDeferredR) == self.count
        assert len(taxFreeR) == self.count

        for i in range(self.count):
            assert len(taxableR[i]) == 3 and \
                   len(taxDeferredR[i]) == 3 and \
                   len(taxFreeR[i]) == 3
            assert (abs(sum(taxableR[i]) - 100) < 0.0000001 and
                    abs(sum(taxDeferredR[i]) - 100) < 0.0000001 and
                    abs(sum(taxFreeR[i]) - 100) < 0.0000001)

    def readContributions(self, filename):
        '''
        Provide the name of the file containing the financial events over the
        anticipated life span determined by the assumed longevity.
        File can be an excel, or odt file with one tab named after
        each spouse and must have the following column headers:

                'year',
                'anticipated income',
                'ctrb taxable',
                'ctrb 401k',
                'ctrb Roth 401k',
                'ctrb IRA',
                'ctrb Roth IRA',
                'Roth X',
                'big ticket items'

        in any order. A template is provided as an example.
        '''
        self.names, self.timeLists = readTimeLists(filename, self.count)

        checkTimeLists(self.names, self.timeLists, self.horizons)

    def setSpousalSplit(self, split):
        '''
        Specify ration of withdrawals between spousal accounts.
        Default value is 'auto'.
        '''
        if type(split) == float:
            assert (0 <= split and split <= 1)

        if split != 'auto':
            u.xprint('Unknown split keyword:', split)

        self.split = split
        u.vprint('Using spousal split of', split)

    def getSplit(self, oldsplit, amount, n, surviving):
        '''
        Calculate auto split based on spousal asset ratio.
        '''
        if surviving == 1:
            return oldsplit
        elif isinstance(self.split, float) or isinstance(self.split, int):
            return self.split
        elif self.split == 'auto':
            s1 = np.sum(self.y2accounts['taxable'][n][:])
            s2 = np.sum(self.y2accounts['tax-deferred'][n][:])
            amount = abs(amount)
            if amount <= s1:
                autoSplit = (self.y2accounts['taxable'][n][0]) / (s1+1)
            elif amount <= s2:
                autoSplit = (self.y2accounts['taxable'][n][0] +
                             self.y2accounts['tax-deferred'][n][0]) / (s1+s2+1)
            else:
                s3 = np.sum(self.y2accounts['tax-free'][n][:])
                autoSplit = (self.y2accounts['taxable'][n][0] +
                             self.y2accounts['tax-deferred'][n][0] +
                             self.y2accounts['tax-free'][n][0]) / (s1+s2+s3+1)

            return autoSplit
        else:
            u.xprint('Unknown split keyword:', self.split)

    def setDesiredIncome(self, income, profile):
        '''
        Set the net target income in retirement following the specified
        profile. Profile can be 'flat' or 'smile'.
        '''
        self.target = income
        if profile == 'smile' or profile == 'flat':
            self.profile = profile
        else:
            u.xprint('Unknown profile keyword: ', profile)

        u.vprint('Using desired net income of', d(income),
                 'with', profile, 'profile')

    def setPension(self, amounts, ages):
        '''
        Set amounts of fixed income pensions if any, and age at which
        it will start.
        '''
        assert len(amounts) == self.count
        assert len(ages) == self.count

        self.pensionAmount = amounts
        self.pensionAge = ages
        u.vprint('Setting pension of', amounts, 'at age(s)', ages)

    def computePension(self, n, i):
        '''
        Compute pension (non-indexed) for individual i in plan year n.
        '''
        refAge = self.pensionAge[i]

        if self.y2ages[n][i] >= refAge:
            return self.pensionAmount[i]

        return 0

    def setSocialSecurity(self, amounts, ages):
        '''
        Set amounts of SS income if any, and age at which
        it will start.
        '''
        assert len(amounts) == self.count
        assert len(ages) == self.count

        self.ssecAmount = amounts
        self.ssecAge = ages
        u.vprint('Setting SSA of', amounts, 'at age(s)', ages)

    def computeSS(self, n, rates, i):
        '''
        Compute social security benefits (indexed) given age,
        inflation, and predicted amount.
        '''
        refAge = self.ssecAge[i]

        if self.y2ages[n][i] < refAge:
            return 0

        return tx.inflationAdjusted(self.ssecAmount[i],
                                    self.y2ages[n][i], rates, refAge)

    def transferWealth(self, year, late):
        '''
        Transfer fraction of assets from one spouse to the other.
        Spouses can inherit IRA and Roth and treat them as their own.
        '''
        other = (late+1) % 2
        for key in self.y2accounts.keys():
            self.y2accounts[key][year][other] += \
                    self.beneficiary[late] * self.y2accounts[key][year][late]
            self.y2accounts[key][year][late] = 0

    def run(self):
        '''
        Run a simulation given the underlying assumptions for
        the next horizon years determined by life expectancies.
        '''
        now = datetime.date.today().year

        # Shorter names for class variables:
        ya2taxable = self.y2accounts['taxable']
        ya2taxDef = self.y2accounts['tax-deferred']
        ya2taxFree = self.y2accounts['tax-free']
        filingStatus = self.status

        # All sources of income.
        self.y2source = {'rmd': np.zeros((self.maxHorizon, self.count)),
                         'ssec': np.zeros((self.maxHorizon, self.count)),
                         'pension': np.zeros((self.maxHorizon, self.count)),
                         'div': np.zeros((self.maxHorizon, self.count)),
                         'job': np.zeros((self.maxHorizon, self.count)),
                         'taxable': np.zeros((self.maxHorizon, self.count)),
                         'dist': np.zeros((self.maxHorizon, self.count)),
                         'tax-free': np.zeros((self.maxHorizon, self.count)),
                         'RothX': np.zeros((self.maxHorizon, self.count)),
                         'bti': np.zeros((self.maxHorizon, self.count))
                         }

        # Use shorter names:
        ys2job = self.y2source['job']
        ys2rmd = self.y2source['rmd']
        ys2pension = self.y2source['pension']
        ys2ssec = self.y2source['ssec']
        ys2div = self.y2source['div']
        ys2dist = self.y2source['dist']
        ys2RothX = self.y2source['RothX']
        ys2txfree = self.y2source['tax-free']
        ys2txbl = self.y2source['taxable']
        ys2bti = self.y2source['bti']

        # Beware of names: tax-free income includes
        # distributions from taxable, tax-free account,
        # and part of SS income.
        self.yincome = {'RothX': np.zeros((self.maxHorizon)),
                        'gross': np.zeros((self.maxHorizon)),
                        'taxes': np.zeros((self.maxHorizon)),
                        'irmaa': np.zeros((self.maxHorizon)),
                        'net': np.zeros((self.maxHorizon)),
                        'target': np.zeros((self.maxHorizon)),
                        'taxable': np.zeros((self.maxHorizon)),
                        'tax-free': np.zeros((self.maxHorizon))
                        }

        # Shorter names:
        yRothX = self.yincome['RothX']
        yincomeTax = self.yincome['taxes']
        yirmaa = self.yincome['irmaa']
        ytargetIncome = self.yincome['target']
        ynetIncome = self.yincome['net']
        ygrossIncome = self.yincome['gross']
        ytaxfreeIncome = self.yincome['tax-free']
        ytaxableIncome = self.yincome['taxable']

        # Assumptions/methods:
        # - Simulation starts at the beginning of the year;
        # - Roth conversions are made at the beginning of the year;
        # - Contributions are made in the middle of the year;
        # - Withdrawals are made in the middle of the year;
        # - Taxes are paid in their current year (as estmated tax);
        # - All balances are calculated for beginning of next year.

        target = self.target

        # For each year ahead:
        u.vprint('Computing next', self.maxHorizon,
                 'years for', [self.names[i] for i in range(self.count)])

        # Keep track of surviving spouses.
        surviving = self.count
        # Use 1 as default for deposit and withdrawal ratios.
        wdrlRatio = 1
        depRatio = 1
        for n in range(0, self.maxHorizon):
            u.vprint('-------', self.yyear[n],
                     ' -----------------------------------------------')

            # Tracker for taxable distribution related to big items.
            btiEvent = 0
            for i in 0, 1:
                # Is the nth year pass i's life horizon?
                if n > self.horizons[i]:
                    u.vprint('Skipping', self.names[i], 'in', self.yyear[n])
                    continue

                # Perform requested Roth conversions early in the year.
                # Keep Roth conversions separately as they are not true income
                # but are taxable events.
                # We will add them separately to taxable income we call gross.
                reqRoth = self.timeLists[i]['Roth X'][n]
                tmp = min(reqRoth, ya2taxDef[n][i])
                if tmp != reqRoth:
                    u.vprint('WARNING: Insufficient funds for Roth conversion for',
                             self.names[i], 'in', self.yyear[n])
                    u.vprint('\tRequested:', d(reqRoth), 'Performed:', d(tmp))
                if tmp > 0:
                    u.vprint(self.names[i], 'requested Roth conversion:',
                             d(reqRoth), ' performed:', d(tmp))
                    ya2taxDef[n][i] -= tmp
                    ya2taxFree[n][i] += tmp
                    ys2RothX[n][i] = tmp
                    yRothX[n] += tmp

                # Add anticipated income for the year.
                tmp = self.timeLists[i]['anticipated income'][n]
                if tmp > 0:
                    u.vprint(self.names[i], 'reported income of', d(tmp))
                    ys2job[n][i] += tmp
                    ytaxableIncome[n] += tmp

                # Add contributions and growth to taxable account.
                # Year-end growth assumes contributions are in midyear.
                # Use += for next year to avoid overwriting inheritance.
                # Else, arrays were initialized to zero.
                ctrb = self.timeLists[i]['ctrb taxable'][n]
                growth = (ya2taxable[n][i] + 0.5*ctrb) * \
                    pfReturn(self.y2assetRatios['taxable'], self.rates, n, i)
                ys2div[n][i] = min(0, growth)
                ya2taxable[n+1][i] += ya2taxable[n][i] + ctrb + growth
                ytaxableIncome[n] += min(0, growth)
                u.vprint(self.names[i], 'Taxable account growth:',
                         d(ya2taxable[n][i]), '->', d(ya2taxable[n+1][i]))

                # Same for tax-deferred, including RMDs on year-end balance.
                ctrb = self.timeLists[i]['ctrb 401k'][n] + \
                    self.timeLists[i]['ctrb IRA'][n]
                growth = (ya2taxDef[n][i] + 0.5*ctrb) * \
                    pfReturn(self.y2assetRatios['tax-deferred'],
                             self.rates, n, i)
                ya2taxDef[n+1][i] += ya2taxDef[n][i] + ctrb + growth
                u.vprint(self.names[i], 'Tax-deferred account growth:',
                         d(ya2taxDef[n][i]), '->', d(ya2taxDef[n+1][i]))
                rmd = ya2taxDef[n+1][i] * \
                    tx.rmdFraction(self.yyear[n], self.yob[i])
                ya2taxDef[n+1][i] -= rmd
                ys2rmd[n][i] = rmd
                ytaxableIncome[n] += rmd

                # And contributions to tax-free accounts:
                ctrb = (self.timeLists[i]['ctrb Roth 401k'][n] +
                        self.timeLists[i]['ctrb Roth IRA'][n])

                growth = (ya2taxFree[n][i] + 0.5*ctrb) * \
                    pfReturn(self.y2assetRatios['tax-free'], self.rates, n, i)
                ya2taxFree[n+1][i] += ya2taxFree[n][i] + ctrb + growth
                u.vprint(self.names[i], 'Tax-free account growth:',
                         d(ya2taxFree[n][i]), '->', d(ya2taxFree[n+1][i]))

                # Compute fixed income for this year:
                ys2pension[n][i] = self.computePension(n, i)
                ytaxableIncome[n] += ys2pension[n][i]
                ys2ssec[n][i] = self.computeSS(n, self.rates, i)
                # Assume our revenues are such that 85% of SS is taxable.
                # Fix if needs arises.
                ytaxfreeIncome[n] += 0.15*ys2ssec[n][i]
                ytaxableIncome[n] += 0.85*ys2ssec[n][i]

                # Big-ticket items can be positive or negative.
                # They do not contribute to income,
                # but withdrawals can be taxable.
                # Take it from the account of bearer: use a split of (i+1)%2.
                bti = self.timeLists[i]['big ticket items'][n]
                if bti != 0:
                    u.vprint(self.names[i],
                             'requested big-ticket item of', d(bti))
                    amounts, total = smartBanking(bti, ya2taxable,
                                                  ya2taxDef, ya2taxFree,
                                                  n+1, (i+1) % 2, self.names)
                    if total != abs(bti):
                        print('WARNING: Insufficient funds for BTI for',
                              self.names[i], 'in', self.yyear[n])
                        print('\tRequested:', d(bti), 'Performed:', d(total))
                        self.success = False

                    # A BTI is not an income unless we created a taxable event
                    # that we track separately (as we do for Roth conversions).
                    btiEvent += amounts['tax-def'][0]
                    ys2dist[n][:] += amounts['tax-def'][1:]
                    ys2txfree[n][:] += amounts['tax-free'][1:]
                    ys2txbl[n][:] += amounts['taxable'][1:]
                    ys2bti[n][i] = math.copysign(total, bti)

            # Compute couple's income needs following profile based on
            # oldest spouse's timeline.
            adjustedTarget = self.target * \
                spendingAdjustment(np.max(self.y2ages[n][:]),
                                   self.profile)
            ytargetIncome[n] = tx.inflationAdjusted(adjustedTarget,
                                                    self.yyear[n],
                                                    self.rates, now)

            gross = ytaxableIncome[n] + yRothX[n] + btiEvent
            estimatedTax = tx.incomeTax(gross, self.yob, filingStatus,
                                        self.yyear[n], self.rates)
            netInc = ytaxfreeIncome[n] + ytaxableIncome[n] - estimatedTax
            gap = netInc - ytargetIncome[n]
            u.vprint('Net income target:', d(ytargetIncome[n]),
                     ' Unadj. net:', d(netInc), ' Delta:', d(gap))
            u.vprint('Taxable:', d(ytaxableIncome[n]),
                     ' Gross:', d(gross), ' Est. Taxes:', d(estimatedTax))

            if gap >= 0:
                if surviving == 2:
                    # Deposit surplus following this year's income ratio.
                    depRatio = (ys2job[n][0] + ys2ssec[n][0] +
                                ys2pension[n][0] + ys2rmd[n][0]) / \
                               (sum(ys2job[n][:] + ys2pension[n][:] +
                                ys2ssec[n][:] + ys2rmd[n][:]) + 1)

                u.vprint('Depositing', d(gap),
                         'in taxable accounts with ratio',
                         '{:.2f}'.format(depRatio))
                smartBanking(gap, ya2taxable, ya2taxDef, ya2taxFree, n+1,
                             depRatio, self.names, True)
                yincomeTax[n] = estimatedTax
                ygrossIncome[n] = gross
                yirmaa[n] = tx.irmaa(gross, filingStatus,
                                     self.yyear[n], self.rates)
                ynetIncome[n] = netInc
            else:
                # Solve amount to withdraw self-consistently.
                # Try at most thirty times.
                # Typically 10 or less iterations are ok.
                # Goal is to reconcile withdrawal in after-tax money,
                # given an existing taxable income.
                withdrawal = gap
                # Adjust withdrawal ratio every year,
                # monitoring surviving spouses.
                wdrlRatio = self.getSplit(wdrlRatio, withdrawal,
                                          n+1, surviving)

                for k in range(30):
                    amounts, total = smartBanking(withdrawal, ya2taxable,
                                                  ya2taxDef, ya2taxFree,
                                                  n+1, wdrlRatio,
                                                  self.names, False)

                    # Zeroth component of amounts countains total.
                    txfree = amounts['taxable'][0] + amounts['tax-free'][0]
                    txbl = amounts['tax-def'][0]
                    totaxblIncome = yRothX[n] + ytaxableIncome[n] + \
                        btiEvent + txbl
                    estimatedTax = tx.incomeTax(totaxblIncome, self.yob,
                                                filingStatus, self.yyear[n],
                                                self.rates)

                    netInc = (txfree + txbl + ytaxfreeIncome[n] +
                              ytaxableIncome[n] - estimatedTax)

                    # No point to loop if we are running out of money.
                    if abs(total - abs(withdrawal)) > 1:
                        print('WARNING: Running out of money!')
                        self.success = False
                        break

                    delta = netInc - ytargetIncome[n]
                    if delta >= -1:
                        u.vprint('Solved with', k, 'iteration(s) and delta of',
                                 d(delta, 2))
                        break

                    withdrawal += delta
                else:
                    u.xprint('Could not converge on withdrawal.')

                # Now commit withdrawal and pay taxes.
                amounts, total = smartBanking(withdrawal, ya2taxable,
                                              ya2taxDef, ya2taxFree,
                                              n+1, wdrlRatio,
                                              self.names, True)

                u.vprint('Performed withdrawal of', d(total),
                         'using split of', '{:.2f}'.format(wdrlRatio))

                txfree = amounts['taxable'][0] + amounts['tax-free'][0]
                txbl = amounts['tax-def'][0]
                ytaxableIncome[n] += txbl
                ys2dist[n][:] += amounts['tax-def'][1:]
                ys2txfree[n][:] += amounts['tax-free'][1:]
                ys2txbl[n][:] += amounts['taxable'][1:]
                ytaxfreeIncome[n] += txfree
                yincomeTax[n] = estimatedTax
                ynetIncome[n] = (ytaxfreeIncome[n] +
                                 ytaxableIncome[n] -
                                 yincomeTax[n])
                gross = ytaxableIncome[n] + yRothX[n] + btiEvent
                ygrossIncome[n] = gross
                yirmaa[n] = tx.irmaa(gross, filingStatus,
                                     self.yyear[n], self.rates)
                u.vprint('\t...of which', d(txbl), 'is taxable.')
                u.vprint('Adj. Income:\n Gross taxable:', d(gross),
                         'Tax bill:', d(yincomeTax[n]),
                         '\n Net:', d(ynetIncome[n]),
                         'Tax free:', d(ytaxfreeIncome[n]))

            # Now check if anyone passed? Then transfer wealth at year-end.
            for j in 0, 1:
                if n == self.horizons[j]:
                    u.vprint(self.names[j], 'has passed.')
                    surviving -= 1
                    if surviving == 0:
                        u.vprint('Both spouses have passed.')
                        return self.yyear, self.y2accounts, \
                            self.y2source, self.yincome

                    xfer = self.beneficiary
                    u.vprint('Transfering', xfer[i], 'fraction of',
                             self.names[i], '\'s wealth to',
                             self.names[(i+1) % 2])
                    self.transferWealth(n+1, j)

                    # Split becomes binary at death of one spouse.
                    wdrlRatio = j
                    depRatio = j
                    filingStatus = 'single'
                    # Reduce target income by 40%.
                    target *= 0.6

        return self.yyear, self.y2accounts, self.y2source, self.yincome

    def plotAccounts(self):
        '''
        Plot values of savings accounts over time.
        '''
        title = 'Savings Balance'
        types = ['taxable', 'tax-deferred', 'tax-free']

        return self.stackPlot(title, self.y2accounts, types, 'upper left')

    def plotSources(self):
        '''
        Plot income over time.
        '''
        title = 'Raw Income Sources'
        types = ['job', 'ssec', 'pension', 'dist', 'rmd', 'RothX',
                 'div', 'taxable', 'tax-free']

        return self.stackPlot(title, self.y2source, types, 'upper left')

    def stackPlot(self, title, accounts, types, location):
        '''
        Core function for stacked plots.
        '''
        import matplotlib.pyplot as plt
        import matplotlib.ticker as tk

        fig, ax = plt.subplots()
        plt.grid(visible='both')
        try:
            get_ipython().__class__.__name__
        except NameError:
            mgr = plt.get_current_fig_manager()
            mgr.window.setGeometry(0, 40, 720, 600)

        accountValues = {}
        for aType in types:
            for i in range(self.count):
                tmp = accounts[aType].transpose()[i]
                if sum(tmp) > 0:
                    accountValues[aType+' '+self.names[i]] = tmp

        ax.stackplot(self.yyear, accountValues.values(),
                     labels=accountValues.keys(), alpha=0.8)
        ax.legend(loc=location, reverse=True)
        # ax.legend(loc=location)
        ax.set_title(title)
        ax.set_xlabel('year')
        ax.set_ylabel('k$')
        ax.get_yaxis().set_major_formatter(
                tk.FuncFormatter(lambda x, p: format(int(x/1000), ',')))

        # plt.show()
        return fig, ax

    def plotNetIncome(self):
        '''
        Plot net income and target over time.
        '''
        title = 'Net Income vs. Target'

        data = {'net': '-', 'target': ':'}
        return self.lineIncomePlot(data, title)

    def lineIncomePlot(self, data, title):
        '''
        Core line plotter function.
        '''
        import matplotlib.pyplot as plt
        import matplotlib.ticker as tk
        from IPython import get_ipython

        fig, ax = plt.subplots()
        plt.grid(visible='both')
        try:
            get_ipython().__class__.__name__
        except NameError:
            mgr = plt.get_current_fig_manager()
            mgr.window.setGeometry(0, 40, 720, 600)

        for aType in data:
            ax.plot(self.yyear, self.yincome[aType],
                    label=aType, ls=data[aType])

        ax.legend(loc='upper left', reverse=True)
        # ax.legend(loc='upper left')
        ax.set_title(title)
        ax.set_xlabel('year')
        ax.set_ylabel('k$')
        ax.get_yaxis().set_major_formatter(
                tk.FuncFormatter(lambda x, p: format(int(x/1000), ',')))

        # plt.show()
        return fig, ax

    def plotTaxableIncome(self):
        '''
        Plot income tax and taxable income over time horizon.
        '''
        import matplotlib.pyplot as plt

        title = 'Taxable Income vs. Tax Brackets'
        data = {'gross': '-'}

        fig, ax = self.lineIncomePlot(data, title)

        myyears = np.array([2022, 2025, 2026, 2052])
        tax2428 = np.array([178000, 220000, 205000, 400000])
        tax3233 = np.array([340000, 405000, 310000, 600000])
        tax35 = np.array([432000, 500000, 580000, 1000000])

        ax.plot(myyears, tax2428, label='24/28%', ls=':')
        ax.plot(myyears, tax3233, label='32/33%', ls=':')
        ax.plot(myyears, tax35, label='35%', ls=':')
        plt.grid(visible='both')
        ax.legend(loc='upper left', reverse=True)
        # ax.legend(loc='upper left')

    def plotTaxes(self):
        '''
        Plot income tax paid over time.
        '''
        title = 'Income Tax and IRMAA'
        data = {'irmaa': '-', 'taxes': '-'}

        fig, ax = self.lineIncomePlot(data, title)

        # plt.show()
        return fig, ax

    def plotRates(self):
        '''
        Plot rate values used over the time horizon.
        '''
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots()
        plt.grid(visible='both')
        try:
            get_ipython().__class__.__name__
        except NameError:
            mgr = plt.get_current_fig_manager()
            mgr.window.setGeometry(800, 40, 720, 600)

        title = 'Return & Inflation Rates ('+str(self.rateMethod)
        if self.rateMethod in ['historical', 'stochastic']:
            title += ' '+str(self.rateFrm)+'-'+str(self.rateTo)
        elif self.rateMethod == 'fixed':
            title += str(self.rateMethod)
        title += ')'
        rateName = ['S&P500 including dividends', 'AA Corporate bonds',
                    '10-y Treasury bonds', 'Inflation']
        ltype = ['-', '-.', ':', '--']
        for i in range(4):
            data = 100*self.rates.transpose()[i]
            label = rateName[i] + ' <' + \
                '{:.2f}'.format(np.mean(data)) + '>'
            ax.plot(self.yyear, data, label=label, ls=ltype[i % 4])

        ax.legend(loc='upper left', reverse=False)
        # ax.legend(loc='upper left')
        ax.set_title(title)
        ax.set_xlabel('year')
        ax.set_ylabel('%')

        # plt.show()
        return fig, ax

    def savePlanXL(self, basename):
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()

        ws = wb.active
        ws.title = 'Income'

        planData = {}
        planData['year'] = self.yyear[:-1]
        planData['target income'] = self.yincome['target'][:-1]
        planData['net income'] = self.yincome['net'][:-1]
        planData['tax bill'] = self.yincome['taxes'][:-1]

        # We need to work by row.
        df = pd.DataFrame(planData)
        for rows in dataframe_to_rows(df, index=False, header=True):
            ws.append(rows)

        # bold = Font(bold=True)
        for cell in ws[1] + ws['A']:
            cell.style = 'Pandas'

        for col in ws.columns:
            column = col[0].column_letter
            # col[0].font = bold
            width = len(str(col[0].value)) + 4
            ws.column_dimensions[column].width = width
            if column != 'A':
                for cell in col:
                    # cell.style = 'Currency [0]'
                    cell.number_format = u'$#,##0_);[Red]($#,##0)'

        for i in range(self.count):
            ws = wb.create_sheet(self.names[i])
            planData = {}
            planData['year'] = self.yyear[:-1]
            planData[self.names[i]+' txbl acc. wrdwl'] = \
                self.y2source['taxable'].transpose()[i][:-1]
            planData[self.names[i]+' RMD'] = \
                self.y2source['rmd'].transpose()[i][:-1]
            planData[self.names[i]+' distribution'] = \
                self.y2source['dist'].transpose()[i][:-1]
            planData[self.names[i]+' Roth conversion'] = \
                self.y2source['RothX'].transpose()[i][:-1]
            planData[self.names[i]+' tax-free wdrwl'] = \
                self.y2source['tax-free'].transpose()[i][:-1]
            planData[self.names[i]+' big-ticket items'] = \
                self.y2source['bti'].transpose()[i][:-1]
            df = pd.DataFrame(planData)
            for rows in dataframe_to_rows(df, index=False, header=True):
                ws.append(rows)

            for cell in ws[1] + ws['A']:
                cell.style = 'Pandas'
            for col in ws.columns:
                column = col[0].column_letter
                # col[0].style = 'Title'
                width = len(str(col[0].value)) + 4
                ws.column_dimensions[column].width = width
                if column != 'A':
                    for cell in col:
                        cell.number_format = u'$#,##0_);[Red]($#,##0)'

        while True:
            try:
                fname = 'plan'+'_'+basename+'.xlsx'
                u.vprint('Saving plan as', fname)
                wb.save(fname)
                break
            except PermissionError:
                print('Failed to save', fname, '. Permission denied.')
                key = input('Try again? [Yn] ')
                if key == 'n':
                    break
            except Exception:
                u.xprint('Unanticipated exception', Exception)

    def savePlanCSV(self, basename):
        import pandas as pd

        planData = {}

        # Start with single entries.
        planData['year'] = self.year
        planData['target income'] = self.yincome['target']
        planData['net income'] = self.yincome['net']
        planData['tax bill'] = self.yincome['taxes']

        for i in range(self.count):
            planData[self.names[i]+' txbl acc. wrdwl'] = \
                self.y2source['taxable'].transpose()[i]
            planData[self.names[i]+' RMD'] = \
                self.y2source['rmd'].transpose()[i]
            planData[self.names[i]+' distribution'] = \
                self.y2source['dist'].transpose()[i]
            planData[self.names[i]+' Roth conversion'] = \
                self.y2source['RothX'].transpose()[i]
            planData[self.names[i]+' tax-free wdrwl'] = \
                self.y2source['tax-free'].transpose()[i]
            planData[self.names[i]+' big-ticket items'] = \
                self.y2source['bti'].transpose()[i]

        df = pd.DataFrame(planData)

        while True:
            try:
                fname = 'plan'+'_'+basename+'.csv'
                df.to_csv(fname)
                # Requires xlwt which is obsolete
                # df.to_excel(fname)
                break
            except PermissionError:
                print('Failed to save', fname, '. Permission denied.')
                key = input('Try again? [Yn] ')
                if key == 'n':
                    break
            except Exception:
                u.xprint('Unanticipated exception', Exception)

    def showAndSave(self, filename=None):
        '''
        Final statement for scripts desiring to save events in excel file.
        '''
        import matplotlib.pyplot as plt
        import os.path as path

        plt.show(block=False)
        plt.pause(0.001)
        while True:
            key = input(
                "[Enter] 'r' to repeat, 's' to save, or 'x' to exit: ")
            if key == 'r':
                plt.close("all")
                break
            elif key == 's':
                if filename is None:
                    filename = path.basename(sys.argv[0][:-3])
                self.savePlanXL(filename)
                break
            elif key == 'x':
                sys.exit(0)

    def estate(self, taxRate):
        '''
        Return estimated post-tax value of total of assets at
        the end of the run in today's $. The tax rate provided
        is used to determine an approximate value and provide
        the tax-deferred portion of the portfolio an after-tax
        value. The inflation rates used during the simulation
        are re-used to bring the net value in today's $.
        '''
        total = sum(self.y2accounts['taxable'][-2][:])
        total += sum(self.y2accounts['tax-free'][-2][:])
        total += (taxRate/100)*sum(self.y2accounts['tax-deferred'][-2][:])

        now = datetime.date.today().year

        return tx.inflationAdjusted(total, now, self.rates, self.yyear[-2])


######################################################################
def d(value, f=0):
    '''
    Return a string formatting number in $ currency.
    '''
    mystr = '{:,.'+str(f)+'f}'
    return '$'+mystr.format(value)


def pfReturn(assetRatios, rates, year, who):
    '''
    Return annual rate of return depending on portfolio asset ratio.
    '''
    return (assetRatios[year][who][0]*rates[year][0] +
            assetRatios[year][who][1]*rates[year][1] +
            assetRatios[year][who][2]*rates[year][2] +
            assetRatios[year][who][3]*rates[year][3])


def age(yob, refYear=0):
    '''
    Return age of individual in year provided. If no year
    is provided, current year will be used.
    '''
    if refYear == 0:
        refYear = datetime.date.today().year

    assert (refYear >= yob)

    return (refYear - yob)


def spendingAdjustment(age, profile='flat'):
    '''
    Return spending profile for age provided.
    Profile can be 'flat' or 'smile'.
    '''
    assert (age <= 100)
    # Return 1 for age before 65 of flat profile.
    if age <= 65 or profile == 'flat':
        return 1.

    if profile == 'smile':
        # From the gogo years to the no-go years.
        table = [1.000, 1.010, 1.015, 1.010, 1.000, 0.993, 0.978,
                 0.960, 0.940, 0.918, 0.895, 0.871, 0.848, 0.825,
                 0.804, 0.785, 0.769, 0.757, 0.748, 0.744, 0.745,
                 0.752, 0.766, 0.787, 0.815, 0.852, 0.899, 0.955,
                 1.021, 1.059, 1.100, 1.121, 1.141, 1.151, 1.161, 1.171]
        return table[age-65]

    u.xprint('In spendingAdjustment: Unknown profile', profile)


def readTimeLists(filename, n):
    '''
    Read listed parameters from an excel spreadsheet through pandas.
    Use one sheet for each individual with the following columns.
    Supports xls, xlsx, xlsm, xlsb, odf, ods, and odt file extensions.
    '''
    import pandas as pd

    # Expected headers in each excel sheet, one per individual.
    timeHorizonItems = ['year',
                        'anticipated income',
                        'ctrb taxable',
                        'ctrb 401k',
                        'ctrb Roth 401k',
                        'ctrb IRA',
                        'ctrb Roth IRA',
                        'Roth X',
                        'big ticket items'
                        ]

    timeLists = []
    names = []
    now = datetime.date.today().year
    # Read all worksheets in memory but only process first n.
    dfDict = pd.read_excel(filename, sheet_name=None)
    i = 0
    for name in dfDict.keys():
        u.vprint('Reading time horizon for', name, '...')
        names.append(name)
        # Only consider lines after this year.
        dfDict[name] = dfDict[name][dfDict[name]['year'] >= now]
        # Replace empty (NaN) cells with 0 value.
        dfDict[name].fillna(0, inplace=True)

        timeLists.append({})
        # Transfer values from dataframe to lists
        for item in timeHorizonItems:
            timeLists[i][item] = dfDict[name][item].tolist()

        i += 1
        if i >= n:
            break

    u.vprint('Successfully read time horizons from file', filename)
    return names, timeLists


def checkTimeLists(names, timeLists, horizons):
    '''
    Make sure that time horizons contain all years up to life expectancy.
    '''
    if len(names) == 2:
        # Verify that both sheets start on the same year.
        if timeLists[0]['year'][0] != timeLists[1]['year'][0]:
            u.xprint('Time horizons not starting on same year.')

    # Verify that year range covers life expectancy for each individual
    now = datetime.date.today().year
    for i in range(0, len(names)):
        yend = now + horizons[i]
        if timeLists[i]['year'][-1] < yend:
            u.xprint('Time horizon for', names[i],
                     'is too short.\n\tIt should end in', yend,
                     'but ends in', timeLists[i]['year'][-1])


def smartBanking(amount, taxable, taxdef, taxfree, year, wdrlRatio,
                 names, commit=True):
    '''
    Deposit/withdraw amount from given accounts. Return dictionary
    of lists itemizing amount taken from taxable, tax-deferred,
    and tax-free accounts, itemized by total and then spousal accounts.
    Total amount from all accounts is second value returned.
    If commit is False, amounts are calculated without changing
    the account values. Withdrawal ratio x controls relative amount
    taken from respective spousal accounts, with reference to first
    spouse (other is 1-x).
    '''
    assert (0 <= wdrlRatio and wdrlRatio <= 1.)
    amounts = {'taxable': [],
               'tax-def': [],
               'tax-free': []
               }
    totAmount = 0
    for i in range(len(names)):
        subAmount = (wdrlRatio - 2*i*wdrlRatio + i)*amount
        itemized = smartBankingSub(subAmount, taxable, taxdef, taxfree,
                                   year, i, names, commit)

        amounts['taxable'].append(itemized[0])
        amounts['tax-def'].append(itemized[1])
        amounts['tax-free'].append(itemized[2])
        totAmount += itemized[3]

    # Store per-account total in first list entry.
    for numlist in amounts.values():
        numlist.insert(0, sum(numlist[:]))

    return amounts, totAmount


def smartBankingSub(amount, taxable, taxdef, taxfree,
                    year, i, names, commit=True):
    '''
    Deposit/withdraw amount from given accounts.
    Return amounts withdrawn from relative accounts:
    taxable, tax-deferred, and tax-free accounts and total withdrawn.
    If commit is False, amounts are calculated without changing account values.
    '''
    if amount == 0:
        return [0, 0, 0, 0]

    # Is this a deposit? Put it in taxable account.
    if amount > 0:
        if commit:
            taxable[year][i] += amount
        return [0, 0, 0, amount]

    # This is a withdrawal. Change sign and start from taxable account.
    withdrawal = abs(amount)
    if withdrawal <= taxable[year][i]:
        if commit:
            taxable[year][i] -= withdrawal
        return [withdrawal, 0, 0, withdrawal]

    portion1 = taxable[year][i]
    if commit:
        taxable[year][i] = 0
    remain = withdrawal - portion1

    # Then go through other accounts. First tax-deferred.
    if remain <= taxdef[year][i]:
        if commit:
            taxdef[year][i] -= remain
        return [portion1, remain, 0, withdrawal]

    portion2 = taxdef[year][i]
    if commit:
        taxdef[year][i] = 0
    remain -= portion2

    # Then tax-free account. Only portion2 is taxable.
    if remain <= taxfree[year][i]:
        if commit:
            taxfree[year][i] -= remain
        return [portion1, portion2, remain, withdrawal]

    portion3 = taxfree[year][i]
    if commit:
        taxfree[year][i] = 0
    remain -= portion3

    if commit:
        print('WARNING: Short withdrawal of', d(amount),
              'in year', year, 'for', names[i])
        print('Missing', d(remain), 'as all accounts were exhausted!')

    return [portion1, portion2, portion3, withdrawal-remain]


